import sys
import openpyxl
from pathlib import Path
from typing import Dict, List
from dataclasses import dataclass
from rdkit import Chem

@dataclass
class HeadgroupPattern:
    name: str
    smarts: str
    lipid_class: str
    fa_positions: List[str]
    description: str = ""


class PatternGenerator:

    @staticmethod
    def load_templates_from_excel(excel_path: str) -> Dict[str, dict]:
        templates = {}
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb["Hoja1"]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 5:
                    continue
                example_num = row[0]
                template_smarts = row[3]
                lipid_type = row[4]
                lipid_subtype = row[5]
                if not template_smarts or not isinstance(template_smarts, str):
                    continue
                if "[G]" not in template_smarts:
                    continue
                template_id = f"template_{example_num}".replace(" ", "_").replace("(", "").replace(")", "")
                templates[template_id] = {
                    "smarts": template_smarts,
                    "lipid_class": lipid_type if lipid_type else "Unknown",
                    "lipid_subtype": lipid_subtype if lipid_subtype else "",
                    "fa_positions": ["N-acyl"] if "SP" in str(lipid_type) else ["sn-1", "sn-2"],
                    "example_num": example_num
                }
        except Exception as e:
            print(f"Error loading templates: {e}", file=sys.stderr)
        return templates

    @staticmethod
    def load_sugars_from_excel(excel_path: str, sheet_name: str = None) -> Dict[str, str]:
        sugars = {}
        try:
            wb = openpyxl.load_workbook(excel_path)
            if sheet_name is None:
                for name in ["Azúcares (G)", "Azúcares", "Sugars", "azucares"]:
                    if name in wb.sheetnames:
                        sheet_name = name
                        break
            if sheet_name not in wb.sheetnames:
                print(f"Warning: Sugar sheet not found. Available: {wb.sheetnames}", file=sys.stderr)
                return {}
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                sugar_name = str(row[0]).strip()
                sugar_smarts = str(row[1]).strip() if len(row) > 1 and row[1] else None
                if sugar_name and sugar_smarts:
                    sugars[sugar_name] = sugar_smarts
            pass
        except Exception as e:
            print(f"Error reading sugars: {e}", file=sys.stderr)
        return sugars

    @staticmethod
    def fa_positions_from_r_count(smarts: str, lipid_type: str) -> List[str]:
        """Derive fa_positions from the number of [R] placeholders in the SMARTS."""
        r_count = smarts.count("[R]")
        if r_count == 0:
            return []
        is_sp = "SP" in str(lipid_type)
        if is_sp:
            return ["N-acyl"] * r_count if r_count == 1 else [f"N-acyl-{i+1}" for i in range(r_count)]
        sn_labels = ["sn-1", "sn-2", "sn-3", "sn-4"]
        if r_count <= len(sn_labels):
            return sn_labels[:r_count]
        return [f"acyl-{i+1}" for i in range(r_count)]

    @staticmethod
    def load_primary_patterns_from_excel(excel_path: str) -> Dict[str, HeadgroupPattern]:
        """Load primary component patterns (rows without [G]) from Excel."""
        patterns = {}
        seen_smarts = set()
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb["Hoja1"]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 5:
                    continue
                example_num = row[0]
                template_smarts = row[3]
                lipid_type = row[4]
                lipid_subtype = row[5]
                if not template_smarts or not isinstance(template_smarts, str):
                    continue
                if "[G]" in template_smarts:
                    continue
                # Replace [R] with generic carbon for matching, then normalise
                # aromatic rings (Kekulé form in Excel → aromatic SMARTS)
                matchable_smarts = template_smarts.replace("[R]", "[#6]")
                matchable_smarts = PatternGenerator.normalise_kekulised_smarts(matchable_smarts)
                if matchable_smarts in seen_smarts:
                    continue
                seen_smarts.add(matchable_smarts)
                fa_pos = PatternGenerator.fa_positions_from_r_count(
                    template_smarts, lipid_type
                )
                safe_id = (f"excel_primary_{example_num}"
                           .replace(" ", "_").replace("(", "").replace(")", "")
                           .replace("/", "_").replace("-", "_"))
                # Ensure unique key
                base_id = safe_id
                counter = 1
                while safe_id in patterns:
                    safe_id = f"{base_id}_{counter}"
                    counter += 1
                lipid_subtype_str = str(lipid_subtype).strip() if lipid_subtype else ""
                pattern_name = lipid_subtype_str if lipid_subtype_str else str(example_num)
                patterns[safe_id] = HeadgroupPattern(
                    name=pattern_name,
                    smarts=matchable_smarts,
                    lipid_class=str(lipid_type) if lipid_type else "Unknown",
                    fa_positions=fa_pos,
                    description=f"Excel primary pattern row {row_idx} (ex {example_num})"
                )
        except Exception as e:
            print(f"Error loading primary patterns: {e}", file=sys.stderr)
        return patterns

    @staticmethod
    def normalise_kekulised_smarts(smarts: str) -> str:
        """
        Round-trip through SMILES to convert Kekulé ring notation (C=C-C=C) to
        aromatic SMARTS (c:c) where RDKit perceives aromaticity.
        """
        if "[*]" in smarts:
            return smarts  # wildcard patterns don't need aromatic normalisation
        as_smiles = smarts.replace("[#6]", "[At]")
        mol = Chem.MolFromSmiles(as_smiles)
        if mol is None:
            return smarts  # fallback: original unchanged
        canonical = Chem.MolToSmarts(mol)
        return canonical.replace("[At]", "[#6]")

    @staticmethod
    def normalise_sugar_smarts(sugar_smarts: str) -> str:
        s = sugar_smarts
        if s.startswith("[G]O"):
            s = s[4:]  # drop "[G]O" – the template's O serves as glycosidic O
        elif s.startswith("[G]"):
            s = s[3:]
        return s

    @staticmethod
    def generate_patterns(templates: Dict, sugars: Dict[str, str]) -> Dict[str, HeadgroupPattern]:
        #Generate all pattern combinations: templates x sugars.
        patterns = {}
        for template_id, template_info in templates.items():
            template_smarts = template_info["smarts"]
            if "[G]" not in template_smarts:
                continue
            for sugar_name, sugar_smarts in sugars.items():
                clean_sugar = PatternGenerator.normalise_sugar_smarts(sugar_smarts)
                complete_smarts = (
                    template_smarts
                    .replace("[G]", clean_sugar, 1)    # specific sugar at primary site
                    .replace("[G]", "[#6]")             # generic carbon at remaining sites
                    .replace("[R]", "[#6]")             # generic carbon at FA attachment
                )
                safe_sugar_name = (sugar_name
                    .replace(" ", "_").replace("(", "").replace(")", "").replace("‑", "_"))
                pattern_id = f"{template_id}_{safe_sugar_name}"
                lipid_subtype = template_info.get("lipid_subtype", "")
                pattern_name = (
                    f"{lipid_subtype} ({sugar_name})" if lipid_subtype
                    else f"Example {template_info['example_num']} ({sugar_name})"
                )
                patterns[pattern_id] = HeadgroupPattern(
                    name=pattern_name,
                    smarts=complete_smarts,
                    lipid_class=template_info["lipid_class"],
                    fa_positions=template_info["fa_positions"],
                    description=f"Auto-generated from template {template_info['example_num']} + {sugar_name}"
                )
        return patterns


def build_combined_patterns(manual_patterns: Dict[str, HeadgroupPattern],
                            excel_path: str = None) -> Dict[str, HeadgroupPattern]:
    all_patterns = dict(manual_patterns)
    if excel_path and Path(excel_path).exists():
        try:
            # Load primary patterns (no [G]) from Excel
            primary = PatternGenerator.load_primary_patterns_from_excel(excel_path)
            primary_added = 0
            for pattern_id, pattern in primary.items():
                if pattern_id not in all_patterns:
                    all_patterns[pattern_id] = pattern
                    primary_added += 1

            # Load [G] templates and combine with sugars
            templates = PatternGenerator.load_templates_from_excel(excel_path)
            sugars = PatternGenerator.load_sugars_from_excel(excel_path)
            generated_added = 0
            if templates and sugars:
                generated = PatternGenerator.generate_patterns(templates, sugars)
                for pattern_id, pattern in generated.items():
                    if pattern_id not in all_patterns:
                        all_patterns[pattern_id] = pattern
                        generated_added += 1

            print(
                f"[OK] Patterns Generated",
                file=sys.stderr
            )
        except Exception as e:
            print(f"Pattern generation failed: {e}", file=sys.stderr)
            import traceback
            traceback.print_exc()
            print(f"Using manual patterns only ({len(manual_patterns)} total)", file=sys.stderr)
    else:
        if excel_path:
            print(f"Excel not found: {excel_path}", file=sys.stderr)
        print(f"Using manual patterns only ({len(all_patterns)} total)", file=sys.stderr)
    return all_patterns